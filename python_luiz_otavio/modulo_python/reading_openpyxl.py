from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / "workbook.xlsx"

# carregando planilha
workbook: Workbook = load_workbook(WORKBOOK_PATH)
# worksheet: Worksheet = workbook.active  # type: ignore

# Selecionar Planilha
worksheet: Worksheet = workbook["Minha Planilha"]  # type: ignore

row: tuple[Cell]
for row in worksheet.iter_rows():
    for cell in row:
        print(cell.value, end="\t")
    print()

worksheet["b3"].value = 15

workbook.save(WORKBOOK_PATH)
