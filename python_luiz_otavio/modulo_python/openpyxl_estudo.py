from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
# worksheet: Worksheet = workbook.active  # type: ignore

# Criar Planilha
workbook.create_sheet("Minha Planilha")
# Selecionar Planilha
worksheet: Worksheet = workbook["Minha Planilha"]  # type: ignore
print(workbook.sheetnames)

# Remover Planilha
workbook.remove(workbook['Sheet'])

worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    ["Leonardo", 25, 9.3],
    ["Maria", 14, 4.3],
    ["Luiz", 61, 6.3],
    ["Renata", 23, 7.3],
    ["Bruna", 30, 9.3],
]

# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)


for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)
